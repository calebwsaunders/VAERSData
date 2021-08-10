#! VAERSDataSearch.py - A simple script used to evaluate VAERS data by vaccine type.

import openpyxl
import calendar
import glob
import pyinputplus as pyip
from datetime import date


def get_current_date():
    """Get current date with goal format of: dd MMM YYYY"""
    date_year = date.today().year
    date_month = date.today().month
    month_abbr = calendar.month_abbr[date_month]
    date_day = date.today().day
    return f'{date_day} {month_abbr} {date_year}'

def get_user_input(message):
    """Get input from the user with an individualized message and return the user's input."""
    output = ""
    while True:
        output = input(message)
        print(f"You entered {output}; is this correct?")
        verify = pyip.inputMenu(["Yes", "No"], numbered=True)
        if verify == "Yes":
            break
    return output

def choose_excel_file():
    """Showing the user all of the Excel files in the current working directory and asking them to select one to
    load if they have an ongoing file they are adding to."""
    excel_files_in_directory = glob.glob('*.xlsx')
    print("The following Excel workbooks are in this folder:")
    i = 1
    for file in excel_files_in_directory:
        print(f"{i}: {file}")
        i += 1
    load_current_file = pyip.inputMenu(['Yes', 'No'],
                                       "\nDo you want to pick one of these files to load for the output file?\n",
                                       numbered=True)
    if load_current_file == 'Yes':
        output = pyip.inputMenu(excel_files_in_directory, numbered=True)
        return output
    else:
        return 'None'

def choose_workbook(message):
    """Asking the user to clarify which excel correlates to VAX ID and which to VAX Reports data."""
    files = glob.glob('*xlsx')
    print(message)
    output = pyip.inputMenu(files, numbered=True)
    return output

# Setting up a dictionary to read all the VAX data into.
vax_data_initial = {}

# A variable for the date of the current data.
data_date = get_user_input("What's the date for this data (it's in the name of the zip folder)? ")

# Using the name of the VAX and the Vax type as the key and adding the id numbers as a list of values.
# key(VAX_MANU, VAX_TYPE): value([VAERS_ID])
# Column A: VAERS_ID; Column B: VAX_TYPE; Column C: VAX_MANU(facturer)
vax_data_file = choose_workbook('Which file has the vaccine ID information (Ex: 20XYVAERSVAX)?')
vax_data_wb = openpyxl.load_workbook(vax_data_file)
vax_data_sheet = vax_data_wb.active
for row in range(2, vax_data_sheet.max_row + 1):  # Start at #2 to skip over the header.
    # Get the values of the pertinent cells
    vax_ID = vax_data_sheet[f'A{row}'].value
    vax_name = vax_data_sheet[f'H{row}'].value
    if vax_name in vax_data_initial:
        vax_data_initial[vax_name].append(vax_ID)
    else:
        vax_data_initial[vax_name] = [vax_data_sheet[f'A{row}'].value]

vax_data_wb.close()

# Matching the VAERS_ID with the vaccine type in vax_data_initial.
# Will count if the person was reported to have died.
# Column A: VAERS_ID; Column G: SEX; Column J: DIED; Column U: NUMDAYS
# The data uses a 'Y' to denote "DIED"
vax_reports = {}
vax_reports_file = choose_workbook("Which file has the vaccine report data (Ex: 20XYVAERSDATA)?")
vax_reports_wb = openpyxl.load_workbook(vax_reports_file)
vax_reports_wb_sheet = vax_reports_wb.active
for row in range(2, vax_reports_wb_sheet.max_row + 1):
    vax_report_ID = vax_reports_wb_sheet[f'A{row}'].value
    vax_reported_died = 0
    vax_reported_male = 0
    reported_age = vax_reports_wb_sheet[f'D{row}'].value
    reported_died = vax_reports_wb_sheet[f'J{row}'].value
    reported_male = vax_reports_wb_sheet[f'G{row}'].value
    if reported_died == 'Y':
        vax_reported_died += 1
    if reported_male == 'M':
        vax_reported_male += 1
    vax_reports[vax_report_ID] = [vax_reported_died, vax_reported_male, reported_age]

vax_reports_wb.close()

# Check to see if output Excel already exists.
# Load sheet if exists, else create new file.
chosen_file = choose_excel_file()
output_wb = ""
if chosen_file == 'None':
    output_wb = openpyxl.Workbook()
    chosen_file = get_user_input("What would you like to name the file? ")
else:
    output_wb = openpyxl.load_workbook(chosen_file)

output_wb_sheet = output_wb.create_sheet(index=0, title=data_date)
output_wb_sheet.merge_cells('A1:D1')
output_wb_sheet['A1'] = f"VAERS Data from: {data_date}; Parsed  on: {get_current_date()}"
output_wb_sheet['A2'] = "Vaccine Type"
output_wb_sheet['B2'] = "Number of Reports"
output_wb_sheet['C2'] = "Deaths Reported"
output_wb_sheet['D2'] = "Male Deaths Reported"
output_wb_sheet['E2'] = "Average Age of Reported Death"

# Setting up variables for finding the total deaths and total deaths attributed to COVID vaccines.
total_deaths = 0
total_deaths_covid_vax = 0
vaccine_data_list = []

# A variable for counting the reports that have no age.
number_with_no_reported_age = 0

# Go through and compare each vax ID by manufacturer and type and see how many deaths
# are associated with each.
for vaccine_type in vax_data_initial:
    # Setup variables to capture the data we're looking for.
    total_reported_occurrences = 0
    total_reported_deaths = 0
    total_reported_deaths_for_average = 0
    male_deaths = 0
    total_age_of_reported_deaths = 0
    average_age_at_death = 0
    for vaccine_id in vax_data_initial[vaccine_type]:
        total_reported_occurrences += 1
        if vax_reports[vaccine_id][0] == 1:
            total_reported_deaths += 1

            # Using a try/except block for records that do not have a recorded age.
            try:
                total_age_of_reported_deaths += vax_reports[vaccine_id][2]
                total_reported_deaths_for_average += 1
            except TypeError:
                number_with_no_reported_age += 1
                continue
        if vax_reports[vaccine_id][0] == 1 and vax_reports[vaccine_id][1] == 1:
            male_deaths += 1

    # Tally the total deaths and see if attributed to COVID VAX.
    total_deaths += total_reported_deaths
    try:
        average_age_at_death = total_age_of_reported_deaths / total_reported_deaths_for_average
    except ZeroDivisionError:
        average_age_at_death = 0
    if vaccine_type.__contains__('COVID19'):
        total_deaths_covid_vax += total_reported_deaths

    vaccine_data_list.append([total_reported_deaths, vaccine_type, total_reported_occurrences, male_deaths,
                              average_age_at_death])

row_to_write_to = 3
# Sort the list by reported deaths and write to Excel.
sorted_vaccine_data = sorted(vaccine_data_list, reverse=True)
for vaccine in sorted_vaccine_data:
    # Write values to Excel.
    output_wb_sheet[f'A{row_to_write_to}'] = vaccine[1]
    output_wb_sheet[f'B{row_to_write_to}'] = vaccine[2]
    output_wb_sheet[f'C{row_to_write_to}'] = vaccine[0]
    output_wb_sheet[f'D{row_to_write_to}'] = vaccine[3]
    output_wb_sheet[f'E{row_to_write_to}'] = int(vaccine[4])
    row_to_write_to += 1

output_wb_sheet['G2'] = "Total Deaths"
output_wb_sheet['G3'] = total_deaths
output_wb_sheet['G5'] = "COVID19 Vaccine Deaths"
output_wb_sheet['G6'] = total_deaths_covid_vax
output_wb_sheet['G8'] = "Non-COVID Vaccine Deaths"
output_wb_sheet['G9'] = total_deaths - total_deaths_covid_vax
output_wb_sheet['G11'] = "COVID19 Vaccine Deaths / Total Deaths"
if total_deaths == 0:
    output_wb_sheet['G12'] = 0
else:
    output_wb_sheet['G12'] = "{:.2%}".format(total_deaths_covid_vax / total_deaths)
output_wb_sheet['G14'] = "Number with no Age"
output_wb_sheet['G15'] = number_with_no_reported_age

# Clean up the spreadsheet.
sheets = output_wb.sheetnames
if 'Sheet' in sheets:
    del output_wb['Sheet']

if chosen_file.endswith('.xlsx'):
    output_wb.save(chosen_file)
    output_wb.close()
else:
    output_wb.save(f'{chosen_file}.xlsx')
    output_wb.close()